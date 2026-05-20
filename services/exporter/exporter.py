"""
lighting-ai/services/exporter/exporter.py
M6 DXF export, M7 Excel BOM, M8 PDF documentation.
"""
from __future__ import annotations
import datetime
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional
import sys

import ezdxf
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from jinja2 import Template

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from config import EXPORTS_DIR
from services.placer.real_placer import PlacementResult, PlacedLuminaire
from services.classifier.room_classifier_real import ClassifiedPlan

LUMI_LAYER="LUMINAIRES"; ZONE_LAYER="ZONES"
COLOR_A=6; COLOR_B=1; COLOR_ZONE=3

# ── DXF ───────────────────────────────────────────────────────────────────────
def _ensure_block(doc,product_code,cutout_mm,beam_angle,color):
    name=product_code.replace('-','_').replace('+','P')[:40]
    if name in doc.blocks: return name
    blk=doc.blocks.new(name=name); r=cutout_mm/2
    blk.add_circle((0,0),r,dxfattribs={"layer":"0","color":color})
    blk.add_circle((0,0),r*0.55,dxfattribs={"layer":"0","color":color})
    blk.add_line((-r*0.4,0),(r*0.4,0),dxfattribs={"layer":"0","color":color})
    blk.add_line((0,-r*0.4),(0,r*0.4),dxfattribs={"layer":"0","color":color})
    return name

def export_dwg(result:PlacementResult, classified:ClassifiedPlan,
               source_dxf_path:Optional[str]=None,
               output_path:Optional[str]=None)->Path:
    if source_dxf_path and Path(source_dxf_path).exists():
        try: doc=ezdxf.readfile(source_dxf_path)
        except: doc=ezdxf.new("R2018")
    else:
        doc=ezdxf.new("R2018")
    doc.header["$INSUNITS"]=4
    msp=doc.modelspace()
    for name,color,lw in [(LUMI_LAYER,COLOR_A,35),(ZONE_LAYER,COLOR_ZONE,18)]:
        if name not in doc.layers:
            doc.layers.add(name,dxfattribs={"color":color,"lineweight":lw})
    for lp in result.placed:
        color=COLOR_A if lp.lumi_type=='A' else COLOR_B
        bn=_ensure_block(doc,lp.product_code,lp.cutout_mm,lp.beam_angle_deg,color)
        msp.add_blockref(bn,insert=(lp.x,lp.y),dxfattribs={"layer":LUMI_LAYER,"rotation":lp.rotation,"color":color})
    for zone in classified.zones:
        b=zone.polygon.bounds
        msp.add_lwpolyline([(b[0],b[1]),(b[2],b[1]),(b[2],b[3]),(b[0],b[3])],
                           close=True,dxfattribs={"layer":ZONE_LAYER,"color":COLOR_ZONE})
        msp.add_text(f"{zone.zone_type}\n{zone.area_m2:.1f}m²",
                     dxfattribs={"layer":ZONE_LAYER,"height":300,"color":COLOR_ZONE,
                                 "insert":((b[0]+b[2])/2,(b[1]+b[3])/2)})
    if output_path is None:
        output_path=str(EXPORTS_DIR/f"{Path(result.source_file).stem}_luminaires.dxf")
    doc.saveas(output_path); print(f"DXF → {output_path}"); return Path(output_path)

# ── Excel ─────────────────────────────────────────────────────────────────────
BLUE=PatternFill("solid",fgColor="1F3864"); ALT=PatternFill("solid",fgColor="EBF0FA")
HF=Font(bold=True,color="FFFFFF",size=10,name="Calibri")
BF=Font(size=9,name="Calibri"); TF=Font(bold=True,size=13,name="Calibri",color="1F3864")
TH=Side(style="thin",color="BFBFBF"); CB=Border(left=TH,right=TH,top=TH,bottom=TH)
CC=Alignment(horizontal="center",vertical="center",wrap_text=True)
LC=Alignment(horizontal="left",vertical="center",wrap_text=True)

def _hc(ws,row,col,val,w=None):
    c=ws.cell(row=row,column=col,value=val); c.font=HF; c.fill=BLUE; c.alignment=CC; c.border=CB
    if w: ws.column_dimensions[get_column_letter(col)].width=w
def _dc(ws,row,col,val,alt=False,align=CC):
    c=ws.cell(row=row,column=col,value=val); c.font=BF; c.alignment=align; c.border=CB
    if alt: c.fill=ALT

def export_excel(result:PlacementResult, classified:ClassifiedPlan,
                 project_name:str="Lighting Project",
                 customer:str="Dirk Rossmann GmbH",
                 concept_id:str="rossmann_standard",
                 output_path:Optional[str]=None)->Path:
    wb=openpyxl.Workbook()
    # Cover
    ws=wb.active; ws.title="Cover"; ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=36
    ws['A1']=project_name; ws['A1'].font=TF
    ws['A2']=f"Customer: {customer}"; ws['A2'].font=Font(size=10,color="444444",name="Calibri")
    ws['A3']=f"Concept: {concept_id}"; ws['A3'].font=Font(size=10,color="444444",name="Calibri")
    ws['A4']=f"Generated: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"; ws['A4'].font=Font(size=9,color="888888",name="Calibri",italic=True)
    for i,(lbl,val) in enumerate([
        ("Total luminaires",len(result.placed)),("Type A (15W 40°)",len(result.by_type("A"))),
        ("Type B (20W 60°)",len(result.by_type("B"))),("Total load",f"{result.total_wattage():.0f} W"),
        ("Zones classified",len(classified.zones)),("Grid pitch","1250 mm"),
        ("Ceiling height","3000 mm"),("Product family","MIKA80-E (MAX FRANKE.led)"),
    ],start=6):
        ws.cell(row=i,column=1,value=lbl).font=Font(bold=True,size=10,name="Calibri")
        ws.cell(row=i,column=2,value=val).font=Font(size=10,name="Calibri")
    # BOM
    ws2=wb.create_sheet("Fixture BOM"); ws2.freeze_panes="A3"
    hdrs=[("Zone",18),("Product code",44),("Description",36),("Mfr.",14),
          ("W",8),("lm",8),("Beam°",8),("Qty",7),("Total W",9),("Mounting",16),("IP",7),("Dim.",9)]
    for col,(h,w) in enumerate(hdrs,1): _hc(ws2,2,col,h,w)
    ws2.row_dimensions[2].height=30
    agg=defaultdict(lambda:{"qty":0,"lp":None})
    for lp in result.placed: k=(lp.zone_type,lp.product_code); agg[k]["qty"]+=1; agg[k]["lp"]=lp
    row=3; tot_qty=0; tot_w=0
    for (zt,pc),d in sorted(agg.items()):
        lp=d["lp"]; q=d["qty"]; alt=(row%2==0)
        vals=[zt.replace('_',' ').title(),pc,lp.description,lp.manufacturer,
              lp.wattage,lp.lux_output,int(lp.beam_angle_deg),q,q*lp.wattage,
              lp.mounting_type.replace('_',' '),lp.ip_rating,"Yes" if lp.dimmable else "No"]
        for col,v in enumerate(vals,1):
            _dc(ws2,row,col,v,alt=alt,align=LC if col in(1,2,3,4,10) else CC)
        tot_qty+=q; tot_w+=q*lp.wattage; row+=1
    for col in range(1,len(hdrs)+1):
        c=ws2.cell(row=row,column=col); c.border=CB; c.fill=BLUE; c.font=HF
    ws2.cell(row=row,column=3,value="TOTAL").font=HF; ws2.cell(row=row,column=3).fill=BLUE
    ws2.cell(row=row,column=8,value=tot_qty).font=HF; ws2.cell(row=row,column=8).fill=BLUE
    ws2.cell(row=row,column=9,value=tot_w).font=HF; ws2.cell(row=row,column=9).fill=BLUE
    # Full schedule
    ws3=wb.create_sheet("Full Schedule"); ws3.freeze_panes="A2"
    fh=[("#",5),("Zone",16),("Type",6),("Product code",44),("Description",32),
        ("W",6),("lm",7),("Beam°",7),("X mm",10),("Y mm",10),("Mounting",16),
        ("Grid",7),("Shelf",7),("IP",7),("CRI",6),("CCT K",7)]
    for col,(h,w) in enumerate(fh,1): _hc(ws3,1,col,h,w)
    ws3.row_dimensions[1].height=30
    for idx,lp in enumerate(result.placed,1):
        alt=(idx%2==0)
        vals=[idx,lp.zone_type.replace('_',' ').title(),lp.lumi_type,
              lp.product_code,lp.description,lp.wattage,lp.lux_output,
              int(lp.beam_angle_deg),round(lp.x),round(lp.y),
              lp.mounting_type.replace('_',' '),
              "✓" if lp.grid_snapped else "–","✓" if lp.shelf_aligned else "–",
              lp.ip_rating,lp.cri,lp.cct_k]
        for col,v in enumerate(vals,1):
            _dc(ws3,idx+1,col,v,alt=alt,align=LC if col in(2,4,5,11) else CC)
    if output_path is None:
        output_path=str(EXPORTS_DIR/f"{Path(result.source_file).stem}_fixture_schedule.xlsx")
    wb.save(output_path); print(f"Excel → {output_path}"); return Path(output_path)

# ── PDF ───────────────────────────────────────────────────────────────────────
_TPL="""<!DOCTYPE html><html><head><meta charset="utf-8"><style>
*{box-sizing:border-box;margin:0;padding:0}body{font-family:Arial,sans-serif;font-size:9pt;color:#1a1a2e}
.cov{background:#1F3864;color:#fff;padding:44px}
.cov h1{font-size:18pt;font-weight:700;margin-bottom:6px}.cov .s{font-size:9pt;opacity:.8;margin-top:4px}
.cov .m{font-size:7.5pt;opacity:.55;margin-top:10px}
.sec{padding:18px 44px}h2{font-size:10pt;font-weight:700;color:#1F3864;border-bottom:2px solid #1F3864;
padding-bottom:3px;margin-bottom:8px;margin-top:14px}
.row{display:flex;gap:10px;margin:10px 0}
.st{background:#EBF0FA;border-left:4px solid #1F3864;padding:8px 12px;flex:1}
.st .v{font-size:14pt;font-weight:700;color:#1F3864}.st .l{font-size:7pt;color:#555;margin-top:1px}
table{width:100%;border-collapse:collapse;font-size:8pt;margin-top:6px}
th{background:#1F3864;color:#fff;padding:5px 7px;text-align:left}
td{padding:4px 7px;border-bottom:1px solid #dde4f0}tr:nth-child(even) td{background:#f4f7fd}
.ft{background:#EBF0FA;padding:10px 44px;font-size:7pt;color:#666;border-top:1px solid #d0daea;margin-top:20px}
@page{size:A4;margin:0}
</style></head><body>
<div class="cov"><h1>{{ project_name }}</h1>
<div class="s">Lighting Design Documentation — {{ concept_id }}</div>
<div class="s">{{ customer }}</div>
<div class="m">Generated {{ generated }} · lighting-ai · MAX FRANKE.led MIKA80-E</div></div>
<div class="sec"><h2>Summary</h2><div class="row">
<div class="st"><div class="v">{{ total }}</div><div class="l">Total luminaires</div></div>
<div class="st"><div class="v">{{ tw }} W</div><div class="l">Connected load</div></div>
<div class="st"><div class="v">{{ ta }} A · {{ tb }} B</div><div class="l">Type split (A=15W 40° / B=20W 60°)</div></div>
<div class="st"><div class="v">{{ zones }}</div><div class="l">Zones classified</div></div>
</div></div>
<div class="sec"><h2>Zone Summary</h2><table>
<tr><th>Zone</th><th>Type</th><th>Area m²</th><th>Qty</th><th>Product</th><th>Method</th></tr>
{% for r in zr %}<tr><td>{{ r.label }}</td><td>{{ r.zt }}</td><td>{{ r.area }}</td>
<td>{{ r.qty }}</td><td style="font-size:7pt">{{ r.prod }}</td><td>{{ r.method }}</td></tr>{% endfor %}
</table></div>
<div class="sec"><h2>Fixture Schedule (Summary)</h2><table>
<tr><th>Type</th><th>Product code</th><th>Description</th><th>Qty</th><th>W</th><th>Total W</th><th>Beam</th></tr>
{% for r in br %}<tr><td>{{ r.t }}</td><td style="font-size:7pt">{{ r.c }}</td><td>{{ r.d }}</td>
<td>{{ r.q }}</td><td>{{ r.w }}</td><td><b>{{ r.tw }}</b></td><td>{{ r.b }}°</td></tr>{% endfor %}
<tr style="background:#1F3864;color:#fff;font-weight:700">
<td colspan="3">TOTAL</td><td>{{ total }}</td><td></td><td>{{ tw }} W</td><td></td></tr>
</table></div>
<div class="sec"><h2>Technical Specs</h2><table>
{% for k,v in specs %}<tr><td><b>{{ k }}</b></td><td>{{ v }}</td></tr>{% endfor %}
</table></div>
<div class="ft">Auto-generated by lighting-ai. Subject to designer review. Alle Maße am Bau prüfen!</div>
</body></html>"""

def export_pdf(result:PlacementResult, classified:ClassifiedPlan,
               concept_id:str="rossmann_standard",
               customer:str="Dirk Rossmann GmbH",
               project_name:str="Lighting Project",
               output_path:Optional[str]=None)->Path:
    bom=defaultdict(lambda:{"qty":0,"lp":None})
    for lp in result.placed: bom[lp.product_code]["qty"]+=1; bom[lp.product_code]["lp"]=lp
    br=[{"t":d["lp"].lumi_type,"c":k,"d":d["lp"].description,"q":d["qty"],
         "w":d["lp"].wattage,"tw":d["qty"]*d["lp"].wattage,"b":int(d["lp"].beam_angle_deg)}
        for k,d in sorted(bom.items())]
    zc=Counter(lp.zone_type for lp in result.placed)
    zp={lp.zone_type:lp.product_code for lp in result.placed}
    zr=[{"label":f"Zone {z.polygon_index}","zt":z.zone_type.replace('_',' ').title(),
         "area":round(z.area_m2,1),"qty":zc.get(z.zone_type,0),
         "prod":zp.get(z.zone_type,'—'),"method":z.method} for z in classified.zones]
    specs=[("Grid pitch","1250 mm"),("Ceiling height","3000 mm / Fries 3300 mm"),
           ("Luminaire family","MIKA80-E (MAX FRANKE.led)"),("Cutout dia.","128 mm"),
           ("Outer dia.","140 mm"),("Embed depth","110 mm"),("CCT","3000 K"),
           ("CRI",">90"),("Dimmable","Yes (DV2.5)"),("IP","IP20"),("Tilt/Rotate","35°/355°")]
    html=Template(_TPL).render(project_name=project_name,concept_id=concept_id,customer=customer,
        generated=datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
        total=len(result.placed),tw=round(result.total_wattage()),
        ta=len(result.by_type("A")),tb=len(result.by_type("B")),
        zones=len(classified.zones),br=br,zr=zr,specs=specs)
    stem=Path(result.source_file).stem
    if output_path is None: output_path=str(EXPORTS_DIR/stem)
    try:
        from weasyprint import HTML as WP; out=output_path+".pdf"
        WP(string=html).write_pdf(out); print(f"PDF → {out}"); return Path(out)
    except Exception: pass
    out=output_path+".html"; Path(out).write_text(html,encoding="utf-8")
    print(f"HTML → {out}"); return Path(out)


if __name__=="__main__":
    from services.parser.pdf_parser import RealPlanParser
    from services.classifier.room_classifier_real import RealRoomClassifier
    from services.placer.real_placer import RealLuminairePlacer
    UP=Path("/mnt/user-data/uploads")
    plan=RealPlanParser().parse(UP/"3600_HH_Jungfernstieg_EG_SB_Kassen_20240506.pdf")
    classified=RealRoomClassifier().classify(plan)
    result=RealLuminairePlacer().place_all(plan,classified)
    print(result.summary())
    dwg=export_dwg(result,classified)
    xlsx=export_excel(result,classified,project_name="Rossmann Hamburg Jungfernstieg EG",customer="Dirk Rossmann GmbH")
    pdf=export_pdf(result,classified,project_name="Rossmann Hamburg Jungfernstieg EG",customer="Dirk Rossmann GmbH")
    print(f"\nDXF:  {dwg}\nXLSX: {xlsx}\nPDF:  {pdf}")